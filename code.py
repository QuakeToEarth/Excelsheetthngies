import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheetTitle = sheet.title
sheet.title = 'Paper Sheet'
grid1 = sheet.cell(row = 1, column=1)
grid1.value = "Aliica"
grid2 = sheet.cell(row = 1, column=2)
grid2.value = "Cookie"
wb.save("\\Users\\licia\\Downloads\\demo.xlsx")
# print(sheet.title)